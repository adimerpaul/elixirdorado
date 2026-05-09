#!/usr/bin/env python3
"""
Elixirdorado — Generador de Reportes Excel Profesionales
Uso: echo '<json>' | python3 generar_excel.py
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta de colores ─────────────────────────────────────────────────────────
C_NAVY   = '1E3A5F'
C_BLUE   = '2563EB'
C_BLUE2  = '3B82F6'
C_WHITE  = 'FFFFFF'
C_GRAY   = 'F1F5F9'
C_GRAY2  = 'E2E8F0'
C_GREEN  = 'DCFCE7'
C_GREEN2 = '16A34A'
C_RED    = 'FEE2E2'
C_RED2   = 'DC2626'
C_YELLOW = 'FEF9C3'
C_YEL2   = 'CA8A04'
C_PURPLE = 'F3E8FF'
C_PUR2   = '7C3AED'

def thin_border(color='CBD5E1'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

BORDER      = thin_border()
BORDER_BOLD = thin_border('94A3B8')

def cs(cell, bg=None, fg='111827', bold=False, size=10,
        halign='left', numfmt=None, wrap=False):
    """Apply style to a cell."""
    cell.font      = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    cell.border    = BORDER
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if numfmt:
        cell.number_format = numfmt
    return cell

def header_row(ws, row, headers, widths=None):
    """Writes a dark navy header row."""
    for ci, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=text)
        cs(c, bg=C_NAVY, fg=C_WHITE, bold=True, size=10, halign='center')
        c.border = BORDER_BOLD
    ws.row_dimensions[row].height = 22
    if widths:
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

def write_title_block(ws, sucursal, report_name, fecha, ncols):
    """3-row branded header: navy title, blue subtitle, thin accent."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=f'  ELIXIRDORADO  —  {sucursal.upper()}')
    cs(t, bg=C_NAVY, fg=C_WHITE, bold=True, size=14)
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=f'  {report_name.upper()}   |   Generado: {fecha}')
    cs(s, bg=C_BLUE2, fg=C_WHITE, bold=False, size=10)
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    ws.cell(row=3, column=1).fill = PatternFill('solid', fgColor='D1E3FF')
    ws.row_dimensions[3].height = 5

def write_stats_block(ws, stats, ncols, row_start=4):
    """Write a 2-row colored stats block. stats = [(label, value, bg_color), ...]"""
    cols_used = len(stats) * 2
    col_pairs = [(i*2+1, i*2+2) for i in range(len(stats))]

    for i, (label, val, bg) in enumerate(stats):
        c1, c2 = col_pairs[i]
        ws.merge_cells(start_row=row_start, start_column=c1, end_row=row_start, end_column=c2)
        cell_l = ws.cell(row=row_start, column=c1, value=label)
        cs(cell_l, bg=bg, fg=C_WHITE, bold=False, size=9, halign='center')

        ws.merge_cells(start_row=row_start+1, start_column=c1, end_row=row_start+1, end_column=c2)
        cell_v = ws.cell(row=row_start+1, column=c1, value=val)
        cs(cell_v, bg=bg, fg=C_WHITE, bold=True, size=13, halign='center')

    # Fill remaining cols if any
    for col in range(cols_used+1, ncols+1):
        ws.cell(row=row_start,   column=col).fill = PatternFill('solid', fgColor='F8FAFC')
        ws.cell(row=row_start+1, column=col).fill = PatternFill('solid', fgColor='F8FAFC')

    ws.row_dimensions[row_start].height   = 16
    ws.row_dimensions[row_start+1].height = 26
    # Spacer
    ws.merge_cells(start_row=row_start+2, start_column=1, end_row=row_start+2, end_column=ncols)
    ws.cell(row=row_start+2, column=1).fill = PatternFill('solid', fgColor=C_GRAY2)
    ws.row_dimensions[row_start+2].height = 5

# ── VENTAS ────────────────────────────────────────────────────────────────────
def gen_ventas(wb, data):
    ws = wb.active
    ws.title = 'Ventas'
    sucursal = data.get('sucursal', 'Sucursal')
    fecha    = data.get('fecha', '')
    items    = data.get('items', [])
    ncols    = 7

    write_title_block(ws, sucursal, 'Reporte de Ventas', fecha, ncols)

    # Stats
    completadas = [v for v in items if (v.get('estado') or '').lower() != 'cancelada']
    total_bs  = sum(float(v.get('total', 0) or 0) for v in completadas)
    count     = len(completadas)
    promedio  = total_bs / count if count else 0

    stats = [
        ('Total Ventas',      count,              C_BLUE),
        ('Total Recaudado',   f'Bs. {total_bs:,.2f}', C_GREEN2),
        ('Promedio / Venta',  f'Bs. {promedio:,.2f}', C_PUR2),
    ]
    write_stats_block(ws, stats, ncols)

    # Headers (row 7 = 3 title + 3 stats + 1 spacer)
    HDR_ROW = 7
    headers = ['FECHA', 'FOLIO', 'MÉTODO DE PAGO', 'SUBTOTAL (Bs.)', 'IVA (Bs.)', 'TOTAL (Bs.)', 'ESTADO']
    widths  = [18, 28, 18, 15, 12, 15, 13]
    header_row(ws, HDR_ROW, headers, widths)

    # Data rows
    for i, v in enumerate(items):
        r      = i + HDR_ROW + 1
        estado = (v.get('estado') or '').lower()
        bg     = C_RED if estado == 'cancelada' else (C_GRAY if i % 2 == 0 else C_WHITE)

        cs(ws.cell(row=r, column=1, value=v.get('fecha', '')),            bg=bg)
        cs(ws.cell(row=r, column=2, value=v.get('folio', '')),            bg=bg, fg=C_NAVY, bold=True)
        cs(ws.cell(row=r, column=3, value=(v.get('metodo_pago') or '').capitalize()), bg=bg)

        sub   = float(v.get('subtotal', 0) or 0)
        iva   = float(v.get('iva',      0) or 0)
        total = float(v.get('total',    0) or 0)
        tfg   = C_GREEN2 if estado == 'completada' else (C_RED2 if estado == 'cancelada' else '111827')

        cs(ws.cell(row=r, column=4, value=sub),   bg=bg, halign='right', numfmt='#,##0.00')
        cs(ws.cell(row=r, column=5, value=iva),   bg=bg, halign='right', numfmt='#,##0.00')
        cs(ws.cell(row=r, column=6, value=total), bg=bg, fg=tfg, bold=True, halign='right', numfmt='#,##0.00')

        if estado == 'completada':
            elabel, efg, ebg = 'Completada', C_GREEN2, C_GREEN
        elif estado == 'cancelada':
            elabel, efg, ebg = 'Cancelada',  C_RED2,   C_RED
        else:
            elabel, efg, ebg = estado.capitalize(), C_YEL2, C_YELLOW
        cs(ws.cell(row=r, column=7, value=elabel), bg=ebg, fg=efg, bold=True, halign='center')

        ws.row_dimensions[r].height = 18

    # Totals row
    tot_r = len(items) + HDR_ROW + 1
    total_sub = sum(float(v.get('subtotal', 0) or 0) for v in completadas)
    total_iva = sum(float(v.get('iva',      0) or 0) for v in completadas)
    total_tot = sum(float(v.get('total',    0) or 0) for v in completadas)

    for col in range(1, ncols+1):
        c = ws.cell(row=tot_r, column=col)
        c.fill   = PatternFill('solid', fgColor=C_NAVY)
        c.font   = Font(name='Arial', bold=True, color=C_WHITE, size=10)
        c.border = BORDER_BOLD
        c.alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=tot_r, column=1).value = f'TOTALES  ({len(completadas)} ventas)'
    ws.merge_cells(start_row=tot_r, start_column=1, end_row=tot_r, end_column=3)
    ws.cell(row=tot_r, column=1).alignment = Alignment(horizontal='center', vertical='center')

    for col, val in [(4, total_sub), (5, total_iva), (6, total_tot)]:
        c = ws.cell(row=tot_r, column=col, value=val)
        c.number_format = '#,##0.00'
        c.alignment     = Alignment(horizontal='right', vertical='center')
        c.font          = Font(name='Arial', bold=True,
                               color='90EE90' if col == 6 else C_WHITE, size=10)
        c.fill   = PatternFill('solid', fgColor=C_NAVY)
        c.border = BORDER_BOLD

    ws.row_dimensions[tot_r].height = 24
    ws.freeze_panes = f'A{HDR_ROW + 1}'

    # ── Sheet 2: Resumen por método ───────────────────────────────────────────
    ws2 = wb.create_sheet('Resumen por Método')
    write_title_block(ws2, sucursal, 'Resumen por Método de Pago', fecha, 4)
    header_row(ws2, 5, ['MÉTODO DE PAGO', 'N° VENTAS', 'TOTAL (Bs.)', '% DEL TOTAL'],
               [22, 12, 18, 14])

    metodos = {}
    for v in completadas:
        m = (v.get('metodo_pago') or 'otro').lower()
        metodos.setdefault(m, {'count': 0, 'total': 0.0})
        metodos[m]['count'] += 1
        metodos[m]['total'] += float(v.get('total', 0) or 0)

    color_map = {
        'efectivo':      (C_GREEN,  C_GREEN2),
        'tarjeta':       ('DBEAFE', C_BLUE),
        'transferencia': (C_YELLOW, C_YEL2),
        'qr':            (C_PURPLE, C_PUR2),
    }
    gr = 6
    for m, vals in sorted(metodos.items()):
        pct = vals['total'] / total_tot if total_tot else 0
        bg, fg = color_map.get(m, (C_WHITE, '111827'))
        cs(ws2.cell(row=gr, column=1, value=m.capitalize()),       bg=bg, fg=fg, bold=True)
        cs(ws2.cell(row=gr, column=2, value=vals['count']),         bg=C_WHITE, halign='center')
        cs(ws2.cell(row=gr, column=3, value=vals['total']),         bg=C_WHITE, halign='right', numfmt='#,##0.00')
        cs(ws2.cell(row=gr, column=4, value=pct),                  bg=C_WHITE, halign='right', numfmt='0.0%')
        ws2.row_dimensions[gr].height = 20
        gr += 1

    # Total row
    for col in range(1, 5):
        c = ws2.cell(row=gr, column=col)
        c.fill   = PatternFill('solid', fgColor=C_NAVY)
        c.font   = Font(name='Arial', bold=True, color=C_WHITE, size=10)
        c.border = BORDER_BOLD
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.cell(row=gr, column=1).value = 'TOTAL'
    cs(ws2.cell(row=gr, column=2, value=count),     bg=C_NAVY, fg=C_WHITE, bold=True, halign='center')
    c3 = ws2.cell(row=gr, column=3, value=total_tot)
    c3.fill = PatternFill('solid', fgColor=C_NAVY)
    c3.font = Font(name='Arial', bold=True, color='90EE90', size=10)
    c3.number_format = '#,##0.00'
    c3.alignment = Alignment(horizontal='right', vertical='center')
    c3.border = BORDER_BOLD
    cs(ws2.cell(row=gr, column=4, value=1.0), bg=C_NAVY, fg=C_WHITE, bold=True, halign='right', numfmt='0.0%')
    ws2.row_dimensions[gr].height = 22

# ── INVENTARIO ────────────────────────────────────────────────────────────────
def gen_inventario(wb, data):
    ws = wb.active
    ws.title = 'Inventario General'
    sucursal = data.get('sucursal', 'Sucursal')
    fecha    = data.get('fecha', '')
    items    = data.get('items', [])
    ncols    = 7

    write_title_block(ws, sucursal, 'Reporte de Inventario', fecha, ncols)

    total_prods  = len(items)
    bajo_stock   = sum(1 for p in items
                       if float(p.get('stock_actual', 0) or 0) <= float(p.get('stock_minimo', 0) or 0))
    valor_inv    = sum(float(p.get('precio_venta', 0) or 0) * float(p.get('stock_actual', 0) or 0)
                       for p in items)

    stats = [
        ('Total Productos',   total_prods,           C_BLUE),
        ('Bajo Stock / Agotado', bajo_stock,          C_RED2),
        ('Valor en Inventario',  f'Bs. {valor_inv:,.2f}', C_GREEN2),
    ]
    write_stats_block(ws, stats, ncols)

    HDR_ROW = 7
    headers = ['CÓDIGO', 'NOMBRE DEL PRODUCTO', 'CATEGORÍA',
               'PRECIO VENTA (Bs.)', 'STOCK ACTUAL', 'STOCK MÍNIMO', 'ESTADO']
    widths  = [14, 32, 18, 18, 13, 13, 13]
    header_row(ws, HDR_ROW, headers, widths)

    def write_prod_row(sheet, r, p, alt):
        stock_act = float(p.get('stock_actual', 0) or 0)
        stock_min = float(p.get('stock_minimo', 0) or 0)
        if stock_act <= 0:
            bg, estado, efg = C_RED,    'Sin stock',  C_RED2
        elif stock_act <= stock_min:
            bg, estado, efg = C_YELLOW, 'Bajo stock', C_YEL2
        else:
            bg = C_GRAY if alt else C_WHITE
            estado, efg = 'Normal', C_GREEN2

        cs(sheet.cell(row=r, column=1, value=p.get('codigo_barras') or '-'),   bg=bg, fg='64748B', size=9)
        cs(sheet.cell(row=r, column=2, value=p.get('nombre', '')),              bg=bg, fg=C_NAVY,  bold=True)
        cs(sheet.cell(row=r, column=3, value=p.get('categoria', 'Sin cat.')),   bg=bg)
        cs(sheet.cell(row=r, column=4, value=float(p.get('precio_venta', 0) or 0)),
           bg=bg, fg=C_GREEN2, bold=True, halign='right', numfmt='#,##0.00')
        stock_fg = C_RED2 if stock_act <= stock_min else C_NAVY
        cs(sheet.cell(row=r, column=5, value=int(stock_act)),
           bg=bg, fg=stock_fg, bold=True, halign='center')
        cs(sheet.cell(row=r, column=6, value=int(stock_min)),
           bg=bg, halign='center')
        cs(sheet.cell(row=r, column=7, value=estado),
           bg=bg if estado == 'Normal' else bg, fg=efg, bold=True, halign='center')
        sheet.row_dimensions[r].height = 18

    for i, p in enumerate(items):
        write_prod_row(ws, i + HDR_ROW + 1, p, i % 2 == 0)

    ws.freeze_panes = f'A{HDR_ROW + 1}'

    # ── Sheet 2: Solo bajo stock ──────────────────────────────────────────────
    ws2 = wb.create_sheet('⚠️ Bajo Stock')
    bajos = [p for p in items
             if float(p.get('stock_actual', 0) or 0) <= float(p.get('stock_minimo', 0) or 0)]

    write_title_block(ws2, sucursal, f'ALERTA — {len(bajos)} Productos Bajo Stock', fecha, ncols)
    header_row(ws2, 5, headers, widths)

    for i, p in enumerate(bajos):
        write_prod_row(ws2, i + 6, p, i % 2 == 0)

    ws2.freeze_panes = 'A6'

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    raw  = sys.stdin.buffer.read().decode('utf-8')
    data = json.loads(raw)
    tipo = data.get('tipo', 'ventas')
    out  = data.get('output', '/tmp/reporte.xlsx')

    wb = Workbook()

    if tipo == 'ventas':
        gen_ventas(wb, data)
    elif tipo == 'inventario':
        gen_inventario(wb, data)
    else:
        sys.stderr.write(f'Tipo desconocido: {tipo}\n')
        sys.exit(1)

    wb.save(out)
    sys.stdout.write(f'OK:{out}\n')

if __name__ == '__main__':
    main()
